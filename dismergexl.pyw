import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook

def unmerge_fill_exclude_rows_cols(excel_path, exclude_rows, exclude_col_header):
    wb = load_workbook(excel_path)
    ws = wb.active

    # 병합된 셀을 해제하고 상단 좌측 셀의 값으로 채우기
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row not in exclude_rows:
            top_left_cell_value = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
            ws.unmerge_cells(str(merged_range))
            for row in ws.iter_rows(min_row=merged_range.min_row, max_row=merged_range.max_row,
                                    min_col=merged_range.min_col, max_col=merged_range.max_col):
                for cell in row:
                    cell.value = top_left_cell_value

    # '상품이미지' 열 제외하기
    col_to_exclude = None
    for col in ws.iter_cols(1, ws.max_column):
        if col[0].value == exclude_col_header:
            col_to_exclude = col[0].column
            break

    if col_to_exclude:
        ws.delete_cols(col_to_exclude)

    # 수정된 엑셀 파일 저장
    modified_path = excel_path.replace('.xlsx', '_modified.xlsx')
    wb.save(modified_path)
    return modified_path

def select_file_and_process():
    root = tk.Tk()
    root.withdraw()  # Tkinter 창을 숨김
    file_path = filedialog.askopenfilename()  # 파일 선택 다이얼로그 열기

    if file_path:
        modified_excel_path = unmerge_fill_exclude_rows_cols(file_path, exclude_rows=[1, 2, 3], exclude_col_header="상품이미지")
        print(f'수정된 엑셀 파일이 저장되었습니다: {modified_excel_path}')

select_file_and_process()
